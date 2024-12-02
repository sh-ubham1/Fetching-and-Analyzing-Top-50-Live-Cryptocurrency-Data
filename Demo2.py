import requests
import pandas as pd
import time
from openpyxl import Workbook, load_workbook

# Function to fetch live cryptocurrency data
def fetch_crypto_data():
    url = "https://api.coingecko.com/api/v3/coins/markets"
    params = {
        "vs_currency": "usd",
        "order": "market_cap_desc",
        "per_page": 50,
        "page": 1,
        "sparkline": False
    }
    response = requests.get(url, params=params)
    if response.status_code == 200:
        data = response.json()
        # Extracting required fields
        crypto_data = []
        for coin in data:
            crypto_data.append([
                coin["name"],
                coin["symbol"].upper(),
                coin["current_price"],
                coin["market_cap"],
                coin["total_volume"],
                coin["price_change_percentage_24h"]
            ])
        return crypto_data
    else:
        print("Failed to fetch data. HTTP Status code:", response.status_code)
        return []

# Function to create or update the Excel file
def update_excel(crypto_data, file_name="live_crypto_data.xlsx"):
    headers = [
        "Cryptocurrency Name",
        "Symbol",
        "Current Price (USD)",
        "Market Capitalization",
        "24h Trading Volume",
        "Price Change (24h %)"
    ]
    
    try:
        # Load existing workbook if it exists
        workbook = load_workbook(file_name)
        sheet = workbook.active
        sheet.delete_rows(2, sheet.max_row)  # Clear old data
    except FileNotFoundError:
        # Create a new workbook if the file doesn't exist
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(headers)  # Add headers

    # Append new data
    for row in crypto_data:
        sheet.append(row)
    
    workbook.save(file_name)
    print(f"Data updated in '{file_name}'.")

# Main script for live updates
if __name__ == "__main__":
    while True:
        print("Fetching live cryptocurrency data...")
        data = fetch_crypto_data()
        if data:
            update_excel(data)
        else:
            print("No data fetched. Retrying in 5 minutes...")
        
        # Wait 5 minutes before refreshing
        time.sleep(300)
